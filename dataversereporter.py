#The following program creates an excel file with worksheets showing monthly totals over
#1. downloads by dataverse (high level)
#2. downloads by dataset (lower level)
#3. datatypes loaded
#4. dataset subjects
#5. accounts created by affiliation

import pandas as pd
import json
import numpy as np
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from pymongo import MongoClient
import time
import psycopg2

class SQLreport:
    def __init__(self, params):
        self.level_count = 1
        self.debug = True
        self.start_date = params['start_date']
        self.end_date = params['end_date']
        self.maindir = params['sqldir']
        self.reportdir = params['reportdir']
        if 'reportfile' in params:
             self.reportfile = params['reportfile']
        else:
             self.reportfile = 'dataverse.xlsx'
        self.needs_table_header=True
        self.table_header=[]
        self.conn = psycopg2.connect("host='%s' dbname='%s' user='%s' password='%s'" % (params['pg_host'], params['pg_dbname'], params['pg_dbuser'], params['pg_password']))

        with open("%s/sql/get_objects.sql" % self.maindir) as f:
	    self.get_objects_sql = f.read()
        with open("%s/sql/get_dataverse_name.sql" % self.maindir) as f:
	    self.get_dataverse_name_sql = f.read()
        with open("%s/sql/downloads_by_month.sql" % self.maindir) as f:
	    self.downloads_by_month_sql = f.read()

    ##SHEET 2
        with open("%s/sql/get_datasets.sql" % self.maindir) as f:
	    self.get_datasets_sql = f.read()
        with open("%s/sql/get_dataset_parents.sql" % self.maindir) as f:
	    self.get_dataset_parents_sql = f.read()
        with open("%s/sql/dataset_downloads_by_month.sql" % self.maindir) as f:
	    self.dataset_downloads_by_month_sql = f.read()
        with open("%s/sql/get_dataset_name.sql" % self.maindir) as f:
	    self.get_dataset_name_sql = f.read()
        with open("%s/sql/get_dataset_files.sql" % self.maindir)as f:
	    self.get_dataset_files_sql = f.read()
        with open("%s/sql/get_dataset_status.sql" % self.maindir)as f:
	    self.get_dataset_status_sql = f.read()

    ##SHEET 3
        with open("%s/sql/get_content_types.sql" % self.maindir) as f:
	    self.get_content_types_sql = f.read()

    ##SHEET 4
        with open("%s/sql/get_subjects.sql" % self.maindir) as f:
	    self.get_subjects_sql = f.read()
        with open("%s/sql/get_dataset_subjects.sql" % self.maindir) as f:
	    self.get_dataset_subjects_sql = f.read()

    ##SHEET 5
        with open("%s/sql/get_affiliations.sql" % self.maindir) as f:
	    self.get_affiliations_sql = f.read()

        with open("%s/sql/get_statistics.sql" % self.maindir) as f:
            self.get_statistics_sql = f.read()

        with open("%s/sql/get_users_by_month_affiliations.sql" % self.maindir) as f:
	    self.get_users_by_month_affiliations_sql = f.read()

    #this is a recursive function allowing the drilldown of relationships from whichever level you start with - determined by dataverse_id 
    def getSubDataverses(self, dataverse_id,level):
	#dataverse_id - the id of the root dataverse
	#level - the current level of the root dataverse
	cur = self.conn.cursor()
	cur.execute(self.get_objects_sql.replace("{dataverse_id}", str(dataverse_id)))
	rows = cur.fetchall()
	#
	for row in rows:
		table_row=[]#reset the table row so it can be added with new data
		for r in row:
			table_row.append(r)
		name=""
		#
		if table_row[1]=="Dataverse":
			info = self.getObjectName(table_row[0],self.get_dataverse_name_sql)
			name=info[0]
			array= self.normalizeArray(level,info[1],name, table_row[2])
			self.getDownloadsByMonthResult(table_row[0],array,self.downloads_by_month_sql)
		#check if we should go deeper	
		if level<self.level_count:	
			self.getSubDataverses(table_row[0],level+1)

    def getDatasets(self, route):
        cur = self.conn.cursor()
	cur.execute(self.get_datasets_sql)
	colnames = [desc[0] for desc in cur.description]
	rows = cur.fetchall()
	DEBUG = 0
	data = []
	for row in rows:
		table_row=[]#reset the table row so it can be added with new data
		for r in row:
			table_row.append(r)
		datarow = {}
		for i in range(0, len(row)):
		    v = colnames[i]
		    if DEBUG:
		        print "%s %s %s" % (i, v, row[i])
		    datarow[v] = row[i]

		#print "T %s" % table_row[0]
		try:
			info = getObjectName(table_row[0],get_dataset_name_sql)
		except:
			info = ['', '']
		title=info[0]
		name=info[1]
		datarow['title'] = title
		datarow['name'] = name

		if DEBUG:
		    print "%s %s" % (title, name)
		status =self.getDatasetStatus(table_row[0])
		hierarchy=list(reversed(self.getHierarchy(table_row[0])))
		del hierarchy[0] # remove first item from list
		datarow['status'] = status
		datarow['hierarchy'] = hierarchy
		data.append(datarow)
#		dvnstore.insert(datarow)
		
		#print json.dumps(datarow)
		
		if len(hierarchy) is 0:
			root="Root"
		else:
			root = hierarchy[0]
			del hierarchy[0]
		array=[root,">".join(hierarchy),title,name,"Dataset"]
		array.append(status)
		array.append(table_row[2])

		#
		files_array=self.getDatasetFiles(table_row[0],1)
		try:
			array.append(round(int(sum(files_array))/1024,1))#convert to KB
		except:
			array.append(0)#when NoneType experienced
		if route is "BY_MONTH":
			self.getDownloadsByMonthResult(table_row[0],array,self.dataset_downloads_by_month_sql)
		if route is "BY_CONTENTTYPE":
			self.getDatasetContent(table_row[0],array)	
		if route is "BY_SUBJECT":
			self.getDatasetSubject(table_row[0],array)	

    def getDatasetStatus(self,dataset_id):
        cur = self.conn.cursor()
	cur.execute(self.get_dataset_status_sql.replace("{id}", str(dataset_id)))
	rows = cur.fetchall()
	for r in rows:
		return r[0]


    def getHierarchy(self,object_id):
        cur = self.conn.cursor()
	cur.execute(self.get_dataset_parents_sql.replace("{object_id}", str(object_id)))
	rows = cur.fetchall()
	array=[]
	for r in rows:
		array.append(self.getObjectName(r[0],self.get_dataverse_name_sql)[0])
	return array

    def normalizeArray(self,level,type,name,date):
	_table_row = [None] * (level-1)
	_table_row.append(name)
	#Also append blank cols to the end
	padding = [None] * (self.level_count-level)
	array = np.concatenate((_table_row, padding), axis=0).tolist()
	array.append(type)
	array.append(date)
	return array

    def getObjectName(self,object_id,sql):
        cur = self.conn.cursor()
	cur.execute(sql.replace("{id}", str(object_id)))
	rows = cur.fetchall()
	return [rows[0][0],rows[0][1]]

    def getDownloadsByMonth(self,object_id,sql):
	global start_date,end_date
	return sql.replace("{object_id}", str(object_id)).replace("{start_date}", self.start_date).replace("{end_date}", self.end_date)

    def getDownloadsByMonthResult(self,object_id,_table_row,sql):
	global needs_table_header,ws_row_count,ws_col_start,ws_cols_count
	# global table_header
        cur = self.conn.cursor()
	cur.execute(self.getDownloadsByMonth(object_id,sql))
	rows = cur.fetchall()
	for row in rows:
		#take the first value and add it to the table_header
		#take the second value and add it to the table_row
		if self.needs_table_header:
			self.table_header.append(row[0].strftime("%b")+" - "+row[0].strftime("%Y"))
		r =	row[1]
		if r is None:
   			 r = 0	
		_table_row.append(r)
		#	
	if self.needs_table_header:
		self.needs_table_header=False
		self.table_header.append("Total")#add total col
		self.ws.append(self.table_header)#first line of workbook
	self.ws_row_count+=1
	self.ws_col_end=get_column_letter(self.ws_col_start+self.ws_cols_count)
	_table_row.append("=SUM("+get_column_letter(self.ws_col_start)+str(self.ws_row_count)+":"+self.ws_col_end+str(self.ws_row_count)+")")#add calculated cell
	self.ws.append(_table_row)#second to n line of workbook

    def getDatasetTypes(self):
        cur = self.conn.cursor()
	cur.execute(self.get_content_types_sql)
	rows = cur.fetchall()
	array=[]
	for row in rows:
		array.append(row[0])
	return array

    def getDatasetContent(self,object_id,_table_row):
	#global contenttypes,ws_col_start,ws_row_count
	files_array=self.getDatasetFiles(object_id,0)
	for i in range(len(self.contenttypes)):
		_table_row.append(files_array.count(self.contenttypes[i]))
	self.ws_row_count+=1
	self.ws_col_end=get_column_letter(self.ws_col_start+len(self.contenttypes)-1)

	_table_row.append("=SUM("+get_column_letter(self.ws_col_start)+str(self.ws_row_count)+":"+self.ws_col_end+str(self.ws_row_count)+")")#add calculated cell
	self.ws.append(_table_row)#second to n line of workbook

    def getSubjects(self):
	cur = self.conn.cursor()
	cur.execute(self.get_subjects_sql)
	rows = cur.fetchall()
	array=[]
	for row in rows:
		array.append(row[0])
	return array

    def getDatasetSubject(self,object_id,_table_row):
	#global subjects,ws_col_start,ws_row_count
	subject_array=self.getDatasetSubjects(object_id,0)
	for i in range(len(self.subjects)):
		_table_row.append(int((self.subjects[i] in subject_array) == True))
	self.ws_row_count+=1
	self.ws_col_end=get_column_letter(self.ws_col_start+len(self.subjects)-1)

	_table_row.append("=SUM("+get_column_letter(self.ws_col_start)+str(self.ws_row_count)+":"+self.ws_col_end+str(self.ws_row_count)+")")#add calculated cell
	self.ws.append(_table_row)#second to n line of workbook

    def getDatasetSubjects(self,object_id,return_col):
	cur = self.conn.cursor()
	cur.execute(self.get_dataset_subjects_sql.replace("{id}", str(object_id)))
	rows = cur.fetchall()
	array=[]
	for row in rows:
		array.append(row[return_col])
	return array

    def getDatasetFiles(self,object_id,return_col):
	cur = self.conn.cursor()
	cur.execute(self.get_dataset_files_sql.replace("{id}", str(object_id)))
	rows = cur.fetchall()
	array=[]
	for row in rows:
		array.append(row[return_col])
	return array

    def getAffiliations(self):
	cur = self.conn.cursor()
	cur.execute(self.get_affiliations_sql)
	rows = cur.fetchall()
	array=[]
	for row in rows:
                if self.debug:
		    print "Affiliation: %s" % row[0]
		array.append(row[0])
	return array

    def getStatistics(self):
        cur = self.conn.cursor()
	#print "SSS " % get_statistics_sql
        cur.execute(self.get_statistics_sql)
        rows = cur.fetchall()
	colnames = [desc[0] for desc in cur.description]
        array=[]
        for row in rows:
	 	datarow = {}
                for i in range(0, len(row)):
                    v = colnames[i]
                    datarow[v] = row[i]
		#dvnmeta.insert(datarow)
                array.append(row[0])
        return array

    def getAffiliationsByMonth(self,affiliations):
	global start_date,end_date,ws_col_start,ws_row_count
	#
	cur = self.conn.cursor()
	cur.execute(self.get_users_by_month_affiliations_sql.replace("{start_date}", self.start_date).replace("{end_date}", self.end_date))
	rows = cur.fetchall()

	months=[]
	for row in rows:
		months.append(row[0])
	ordered_dates=sorted(set(months))
	#append months to first row
	for date in ordered_dates:
		self.table_header.append(date.strftime("%b")+" - "+date.strftime("%Y"))
	self.table_header.append("Totals")

	self.ws.append(self.table_header)#first line of workbook
	#
	self.ws_col_end=get_column_letter(self.ws_col_start+ len(ordered_dates)-1)
	#
	for a in affiliations:
		_table_row=[0] * (len(ordered_dates))
		_table_row.insert(0, a)
		for row in rows:
			if str(row[2]) == str(a):
				_table_row[ordered_dates.index(row[0])+1]=row[1]
		#Add totals col
		self.ws_row_count+=1
		_table_row.append("=SUM("+get_column_letter(self.ws_col_start)+str(self.ws_row_count)+":"+self.ws_col_end+str(self.ws_row_count)+")")#add calculated cell
		self.ws.append(_table_row)

    def addWorkSheetFooter(self,start_col,calc_col_count,pad):
	#start_col - the position of the fist calculated col 
	#calc_col_count - the number columns to calculate
	#pad - the number of non-calculated cols
	#
	table_row=["Totals"]
	padding = [None] * pad
	array = np.concatenate((table_row, padding), axis=0).tolist()
	for i in range(0, calc_col_count):
		letter=get_column_letter(start_col+i)
		array.append("=SUM("+letter+"2:"+letter+str(self.ws_row_count)+")")
	self.ws.append(array)#inject last row with totals
	
    def reportmaker(self):
        ####create the workbook
        self.wb = Workbook()
        self.ws = self.wb.active
        print(self.get_dataverse_name_sql)
        if not self.get_dataverse_name_sql:
            return 'Error: sql queries folder not found!'
        self.ws_cols_count=12*3#months of year and number years (depending on date ranges)
#
#SHEET 1 ##########################################
#
        if self.debug:
            print('SHEET 1')
        self.table_header=[]
        self.table_header.append("Top Level")
        self.table_header.append("Category")
        self.table_header.append("Publication Date")
        self.ws.title = "Downloads by Dataverse"#Label the worksheet
        self.ws_row_count=1#increments with each added row - used to calculate the totals
        self.ws_col_start=4#the col to start calculateding from used in the last col sum 
        self.getSubDataverses(1,1)
        self.addWorkSheetFooter(4,self.ws_cols_count+2,2)

#####SHEET 2#######################################
        if self.debug:
            print('SHEET 2')
        self.needs_table_header=True
        self.level_count=1
        self.table_header=[]
        self.table_header.append("Root")
        self.table_header.append("Path")
        self.table_header.append("Title")
        self.table_header.append("Name")
        self.table_header.append("Type")
        self.table_header.append("Status")
        self.table_header.append("Publication Date")
        self.table_header.append("Size (KB)")
        self.wb.create_sheet('Downloads by Dataset')
        self.ws = self.wb["Downloads by Dataset"]
        self.ws_row_count=1
        self.ws_col_start=len(self.table_header)+1

        self.getDatasets("BY_MONTH")
        self.addWorkSheetFooter(self.ws_col_start-1,self.ws_cols_count+3,self.ws_col_start-3)

######SHEET 3
        if self.debug:
            print('SHEET 3')
        self.needs_table_header=True
        self.level_count=1
#
        self.table_header=[]
        self.table_header.append("Root")
        self.table_header.append("Path")
        self.table_header.append("Title")
        self.table_header.append("Name")
        self.table_header.append("Type")
        self.table_header.append("Status")
        self.table_header.append("Publication Date")
        self.table_header.append("Size (KB)")

        self.wb.create_sheet('File Types')
        self.ws = self.wb["File Types"]
        self.ws_row_count=1
        self.ws_col_start=len(self.table_header)+1
        self.contenttypes=self.getDatasetTypes()
        self.table_header= np.concatenate((self.table_header, self.contenttypes), axis=0).tolist()
        self.table_header.append("Totals")
        self.ws.append(self.table_header)#first line of workbook
        self.getDatasets("BY_CONTENTTYPE")
        self.addWorkSheetFooter(self.ws_col_start-1,len(self.contenttypes)+2,self.ws_col_start-3)

######SHEET 4
        if self.debug:
            print('SHEET 4')
        self.needs_table_header=True
        self.level_count=1
#
        self.table_header=[]
        self.table_header.append("Root")
        self.table_header.append("Path")
        self.table_header.append("Title")
        self.table_header.append("Name")
        self.table_header.append("Type")
        self.table_header.append("Status")
        self.table_header.append("Publication Date")
        self.table_header.append("Size (KB)")

        self.wb.create_sheet('Subjects')
        self.ws = self.wb["Subjects"]
        self.ws_row_count=1
        self.ws_col_start=len(self.table_header)+1
        self.subjects=self.getSubjects()
        self.table_header= np.concatenate((self.table_header, self.subjects), axis=0).tolist()
        self.table_header.append("Totals")
        self.ws.append(self.table_header)#first line of workbook
        self.getDatasets("BY_SUBJECT")
        self.addWorkSheetFooter(self.ws_col_start-1,len(self.subjects)+2,self.ws_col_start-3)

###SHEET 5
        if self.debug:
            print('SHEET 5')
        self.needs_table_header=True
        self.ws_row_count=1
        self.ws_col_start=2
        self.table_header=[]
        self.table_header.append("Affiliation")
        self.wb.create_sheet("Users by Affiliations")
        self.ws = self.wb["Users by Affiliations"]
        self.getAffiliationsByMonth(self.getAffiliations())
        self.getStatistics()
        self.addWorkSheetFooter(2,self.ws_cols_count+2,0)

####
        if self.debug:
            print('Saving report')
        reportfile = "%s/%s" % (self.reportdir, self.reportfile)
        self.wb.save(reportfile)
        return reportfile
